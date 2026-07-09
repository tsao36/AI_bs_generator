"""Weekly project/task loading report per engineer (CFE work overview – 2026 tasks tab).

Reads the OneDrive-synced CFE work overview.xlsx, aggregates the '2026 tasks'
sheet by owner, builds a priority-weighted table, generates per-person LLM
comments, and sends the report via Microsoft Graph API.
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import urllib.request
from collections import defaultdict
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from Meeting_agenda_OneNote import (  # type: ignore
    get_graph_token_app_only,
    get_graph_token_delegated_with_secret,
    load_recipients,
    resolve_graph_client_secret,
    send_mail_via_graph,
)

LOG = logging.getLogger("weekly_project_loading_report")

# ── Default Excel path (OneDrive synced SharePoint) ──────────────────────────
_DEFAULT_EXCEL = (
    r"C:\Users\jtsao1\OneDrive - Intel Corporation"
    r"\CFE Team Folder\CFE Work Package Overview\CFE work overview.xlsx"
)
_SHEET_NAME = os.getenv("CFE_TASK_SHEET_NAME", "CFE_tasks").strip() or "CFE_tasks"

# ── Owner name normalisation ─────────────────────────────────────────────────
# Keys are lowercase versions of whatever appears in the spreadsheet.
OWNER_ALIASES: Dict[str, str] = {
    "wesley":              "Wesley Kuo",
    "steven":              "Steven1 Chen",
    "timdaway":            "Timdaway Lai",
    "kj":                  "Kj Fang",
    "brenton":             "Brenton Wu",
    "jonathan":            "Jonathan Tsao",
    "yu-wei":              "Yu-wei Chen",   # covers both "Yu-Wei" and "Yu-wei"
    "henry su":            "Henryx Su",
    "zhiqiang":            "Zhiqiang Cai",
    "charles chu":         "Charles Chu",
    "frank yang":          "Frank Yang",
    "leo chiang":          "Leo Chiang",
    "frank lee":           "Frank Lee",
    "matt chen":           "Matt Chen",
    "juan zou":            "Juan Zou",
    "bingyue sun":         "Bingyue Sun",
    "jackx lee":           "Jackx Lee",
}
_SKIP_OWNERS = {"looking for owner", "account based", "tbd", ""}

# ── Priority mapping ─────────────────────────────────────────────────────────
# Spreadsheet uses numeric values: 3=High, 2=Medium, 1=Low
# Legacy string labels are also accepted for backward compatibility.
PRIO_MAP: Dict[Any, str] = {3: "High", 2: "Medium", 1: "Low",
                             "3": "High", "2": "Medium", "1": "Low",
                             "High": "High", "Medium": "Medium", "Low": "Low",
                             "KTBR": "KTBR"}
PRIO_WEIGHT = {"High": 3, "Medium": 2, "Low": 1, "KTBR": 1}
PRIO_ORDER  = ["High", "Medium", "Low"]  # KTBR removed (no longer in sheet)


# ─────────────────────────────────────────────────────────────────────────────
def _env_str(name: str, default: str = "") -> str:
    return (os.getenv(name) or default).strip() or default


def _normalize_owner(raw: str) -> List[str]:
    """Return list of canonical owner names for a (possibly multi-owner) cell."""
    raw = raw.strip()
    # Split on common separators used in the sheet
    parts = [p.strip() for p in raw.replace("/", ",").split(",") if p.strip()]
    result: List[str] = []
    for part in parts:
        key = part.lower()
        if key in _SKIP_OWNERS:
            continue
        canonical = OWNER_ALIASES.get(key, part)
        result.append(canonical)
    return result


def read_2026_tasks(excel_path: str) -> List[Dict[str, Any]]:
    """Parse the configured task sheet and return one dict per task-owner pair."""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if _SHEET_NAME not in wb.sheetnames:
        available = ", ".join(str(x) for x in wb.sheetnames)
        raise ValueError(f"Sheet '{_SHEET_NAME}' not found in {excel_path}. Available sheets: {available}")
    ws = wb[_SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    header = [str(cell or "").strip().lower() for cell in rows[0]]

    def _column(*names: str, fallback: int) -> int:
        for name in names:
            if name.lower() in header:
                return header.index(name.lower())
        return fallback

    task_idx = _column("tasks", "task", fallback=0)
    year_idx = _column("year", fallback=-1)
    priority_idx = _column("priority", fallback=1)
    owner_idx = _column("owner", fallback=5)
    current_year = datetime.now().year

    tasks: List[Dict[str, Any]] = []
    for row in rows[1:]:          # row[0]=header, skip
        if year_idx >= 0:
            try:
                row_year = int(float(str(row[year_idx] or "").strip()))
            except ValueError:
                continue
            if row_year != current_year:
                continue
        task      = str(row[task_idx] or "").strip()
        prio_raw  = row[priority_idx]
        priority  = PRIO_MAP.get(prio_raw, PRIO_MAP.get(str(prio_raw).strip(), "Low"))
        owner_raw = str(row[owner_idx] or "").strip()
        if not task:
            continue
        for owner in _normalize_owner(owner_raw):
            tasks.append({"task": task, "priority": priority, "owner": owner})
    return tasks


def _build_per_owner(tasks: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Aggregate tasks per owner and return rows sorted by weighted score (desc)."""
    agg: Dict[str, Dict[str, Any]] = defaultdict(
        lambda: {"High": 0, "Medium": 0, "Low": 0, "task_list": []}
    )
    for t in tasks:
        owner = t["owner"]
        prio  = t["priority"]
        if prio in agg[owner]:
            agg[owner][prio] += 1
        agg[owner]["task_list"].append(t["task"])

    rows: List[Dict[str, Any]] = []
    for owner, d in agg.items():
        total = sum(d[p] for p in PRIO_ORDER)
        rows.append({
            "owner":     owner,
            "High":      d["High"],
            "Medium":    d["Medium"],
            "Low":       d["Low"],
            "total":     total,
            "task_list": d["task_list"],
        })
    rows.sort(key=lambda r: (r["High"], r["Medium"], r["total"]), reverse=True)
    return rows


def _format_project_table(rows: List[Dict[str, Any]]) -> str:
    headers  = ["Owner", "High", "Med", "Low", "Total"]
    numeric  = {1, 2, 3, 4}
    col_data = [
        [r["owner"], str(r["High"]), str(r["Medium"]),
         str(r["Low"]), str(r["total"])]
        for r in rows
    ]
    widths = [
        max(len(headers[i]), max((len(c[i]) for c in col_data), default=0))
        for i in range(len(headers))
    ]
    sep = "+-" + "-+-".join("-" * w for w in widths) + "-+"
    hdr = "| " + " | ".join(
        h.rjust(widths[i]) if i in numeric else h.ljust(widths[i])
        for i, h in enumerate(headers)
    ) + " |"
    lines = [sep, hdr, sep]
    for cols in col_data:
        line = "| " + " | ".join(
            cols[i].rjust(widths[i]) if i in numeric else cols[i].ljust(widths[i])
            for i in range(len(headers))
        ) + " |"
        lines.append(line)
    lines.append(sep)
    return "\n".join(lines)


def _generate_zh_project_summary(rows: List[Dict[str, Any]], year: str) -> Optional[str]:
    """Generate a Traditional Chinese paragraph summarising project task assignments."""
    if not rows:
        return None
    base_url = _env_str("LOCAL_LLM_URL", "http://127.0.0.1:11434")
    model    = _env_str("LOCAL_LLM_MODEL", "llama3.1:8b-instruct-q4_K_M")

    total_tasks = sum(r["total"] for r in rows)
    top3 = rows[:3]
    top3_str = "；".join(
        f"{r['owner']} ({r['total']} 項，High {r['High']} / Med {r['Medium']})"
        for r in top3
    )
    high_total = sum(r["High"] for r in rows)

    prompt = (
        f"你是一位無線工程團隊的專案管理分析師。請用繁體中文，以3到5句話，"
        f"用平易近人的語言摘要以下 {year} 年度各工程師專案任務分配現況報告。"
        f"注意：此表格仍在填寫中，數字為部分資料，請勿做過度推論。不要列點，直接寫成段落。\n\n"
        f"統計年度：{year} 年（資料持續更新中）\n"
        f"已分配任務總數：{total_tasks} 項\n"
        f"其中 High Priority 任務：{high_total} 項\n"
        f"任務數最多的前三位：{top3_str}\n"
        f"共統計 {len(rows)} 位工程師\n"
    )

    payload = json.dumps({"model": model, "prompt": prompt, "stream": False}).encode("utf-8")
    try:
        req = urllib.request.Request(
            f"{base_url}/api/generate",
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=60) as resp:
            result = json.loads(resp.read().decode("utf-8"))
        summary = (result.get("response") or "").strip()
        if summary:
            LOG.info("Generated Traditional Chinese project summary (%d chars).", len(summary))
            return summary
    except Exception as exc:
        LOG.warning("Local LLM unavailable for zh summary: %s", exc)
    return None


def _generate_project_comments(rows: List[Dict[str, Any]], year: str) -> Dict[str, str]:
    """Call local LLM for a one-line comment per owner based on task loading."""
    if not rows:
        return {}
    base_url = _env_str("LOCAL_LLM_URL", "http://127.0.0.1:11434")
    model    = _env_str("LOCAL_LLM_MODEL", "llama3.1:8b-instruct-q4_K_M")

    entries = [
        f"  {r['owner']}: High={r['High']}, Med={r['Medium']}, Low={r['Low']}, "
        f"Total={r['total']}"
        for r in rows
    ]
    prompt = (
        f"You are a wireless engineering team lead reviewing {year} project task assignments.\n"
        f"The task table is still being filled in, so totals are partial and not final.\n"
        f"For EACH person listed, write exactly ONE concise English sentence (max 20 words) "
        f"acknowledging their current assigned tasks without making workload judgements.\n"
        f"Output format (one line per person, exactly):\n"
        f"Name: <comment>\n\n"
        f"Data (columns: High/Med/Low priority task counts, Total):\n"
        + "\n".join(entries) + "\n\n"
        f"Do not add extra lines, headers, or explanations."
    )

    payload = json.dumps({"model": model, "prompt": prompt, "stream": False}).encode("utf-8")
    try:
        req = urllib.request.Request(
            f"{base_url}/api/generate",
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=120) as resp:
            result = json.loads(resp.read().decode("utf-8"))
        text = (result.get("response") or "").strip()
        comments: Dict[str, str] = {}
        for line in text.splitlines():
            line = line.strip()
            if ":" in line:
                key, _, val = line.partition(":")
                key = key.strip(); val = val.strip()
                if key and val:
                    comments[key] = val
        LOG.info("Generated project LLM comments for %d owners.", len(comments))
        return comments
    except Exception as exc:
        LOG.warning("Local LLM unavailable for project comments: %s", exc)
        return {}


def _build_body(rows: List[Dict[str, Any]], excel_path: str,
                as_of: str, zh_summary: Optional[str],
                comments: Optional[Dict[str, str]]) -> str:
    year        = as_of[:4]
    total_tasks = sum(r["total"] for r in rows)
    lines: List[str] = []

    if zh_summary:
        lines.append(f"[{year} 年度各工程師專案任務分配摘要（繁體中文）]")
        lines.append(zh_summary)
        lines.append("")
        lines.append("=" * 60)
        lines.append("")

    lines.append(f"{year} Project / Task Loading per Engineer")
    lines.append("=" * 60)
    lines.append(f"Generated at : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"Source file  : {excel_path}")
    lines.append(f"Source sheet : {_SHEET_NAME}")
    lines.append(f"Engineers    : {len(rows)}")
    lines.append(f"Total tasks  : {total_tasks}  (partial — table still being updated)")
    lines.append("")

    lines.append(_format_project_table(rows))
    lines.append("")

    lines.append("Column definitions:")
    lines.append("  High   = Priority 3 tasks")
    lines.append("  Med    = Priority 2 tasks")
    lines.append("  Low    = Priority 1 tasks")
    lines.append("  Total  = High + Med + Low")
    lines.append("  Note: tasks marked 'Account based' or 'Looking for owner' are excluded")
    lines.append("  Note: table is still being filled in — counts are partial and not final")
    lines.append("")

    if comments:
        lines.append("=" * 60)
        lines.append("AI Feedback per Engineer  (generated by local LLM)")
        lines.append("=" * 60)
        lines.append("  [Note: This section is still a work in progress and will be improved in future iterations.]")
        lines.append("")
        for r in rows:
            comment = comments.get(r["owner"], "")
            if comment:
                lines.append(f"  {r['owner']}: {comment}")
        lines.append("")

    lines.append("This email was auto-generated by run_weekly_current_yearly_issue_count.bat")
    return "\n".join(lines)


def _build_subject(rows: List[Dict[str, Any]], as_of: str) -> str:
    year  = as_of[:4]
    total = sum(r["total"] for r in rows)
    return (
        f"[{year} Project Loading] {as_of} | "
        f"{len(rows)} engineers | {total} tasks assigned (partial)"
    )


def _setup_logging(level: str = "INFO") -> None:
    logging.basicConfig(level=getattr(logging, level.upper(), logging.INFO),
                        format="%(asctime)s [%(levelname)s] %(message)s")
    LOG.setLevel(getattr(logging, level.upper(), logging.INFO))


def _get_token(graph_auth_mode: str) -> Tuple[str, str]:
    secret = resolve_graph_client_secret()
    mode   = (graph_auth_mode or _env_str("GRAPH_AUTH_MODE", "app")).lower()
    if mode == "delegated":
        token = get_graph_token_delegated_with_secret(secret, scopes=["Mail.Send"])
        return token, ""
    sender_upn = _env_str("GRAPH_SENDER_UPN")
    if not sender_upn:
        raise RuntimeError("GRAPH_SENDER_UPN is required for app mode.")
    token = get_graph_token_app_only(
        secret, scopes=["https://graph.microsoft.com/.default"]
    )
    return token, sender_upn


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Send weekly project task loading report.")
    parser.add_argument("--excel", default=_env_str("CFE_EXCEL_PATH", _DEFAULT_EXCEL),
                        help="Path to CFE work overview.xlsx")
    parser.add_argument("--recipients", default="recipients.json")
    parser.add_argument("--extra-to", default="",
                        help="Additional comma-separated To addresses.")
    parser.add_argument("--graph-auth-mode", choices=["app", "delegated"],
                        default=_env_str("GRAPH_AUTH_MODE", "delegated"))
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--log-level", default="INFO")
    args = parser.parse_args()

    _setup_logging(args.log_level)

    as_of = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    year  = as_of[:4]

    LOG.info("Reading tasks from: %s", args.excel)
    tasks = read_2026_tasks(args.excel)
    LOG.info("Loaded %d task-owner entries.", len(tasks))

    rows = _build_per_owner(tasks)
    LOG.info("Aggregated %d owners.", len(rows))

    comments   = _generate_project_comments(rows, year)
    zh_summary = _generate_zh_project_summary(rows, year)
    body       = _build_body(rows, args.excel, as_of, zh_summary, comments)
    subject  = _build_subject(rows, as_of)

    to_list, cc_list = load_recipients(args.recipients)
    if args.extra_to:
        extras = [x.strip() for x in args.extra_to.split(",") if x.strip()]
        seen   = {x.lower() for x in to_list}
        for addr in extras:
            if addr.lower() not in seen:
                to_list.append(addr)
                seen.add(addr.lower())

    if not to_list:
        LOG.error("No recipients configured.")
        return 1

    if args.dry_run:
        print("[DRY RUN] To:", to_list)
        print("[DRY RUN] Cc:", cc_list)
        print("[DRY RUN] Subject:", subject)
        print("\n--- Body Preview ---\n")
        print(body)
        return 0

    token, sender_upn = _get_token(args.graph_auth_mode)
    import html as _html
    html_body = (
        "<html><body>"
        "<pre style='font-family: Consolas, Courier New, monospace; font-size: 13px;'>"
        + _html.escape(body)
        + "</pre></body></html>"
    )
    send_mail_via_graph(
        token=token,
        subject=subject,
        body_text=html_body,
        to_addrs=to_list,
        cc_addrs=cc_list or None,
        save_to_sent_items=True,
        content_type="HTML",
        sender_upn=sender_upn or None,
    )
    LOG.info("Email sent to %d recipients.", len(to_list))
    if cc_list:
        LOG.info("Cc: %s", ", ".join(cc_list))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
