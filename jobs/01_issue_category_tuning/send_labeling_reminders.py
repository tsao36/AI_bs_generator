from __future__ import annotations

import argparse
import csv
import glob
import json
import logging
import os
import re
import shutil
import sys
import tempfile
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime
from typing import Dict, List

from Meeting_agenda_OneNote import (  # type: ignore
    get_graph_token_app_only,
    get_graph_token_delegated_with_secret,
    load_recipients,
    resolve_graph_client_secret,
    send_mail_via_graph,
)

LOG = logging.getLogger("labeling_reminder")
GRAPH_BASE = "https://graph.microsoft.com/v1.0"


def _setup_logging(level: str) -> None:
    numeric_level = getattr(logging, str(level or "INFO").upper(), logging.INFO)
    logging.basicConfig(level=numeric_level, format="%(asctime)s [%(levelname)s] %(message)s")
    LOG.setLevel(numeric_level)


def _find_latest_weekly_dir(tuning_root: str) -> str:
    candidates = [p for p in glob.glob(os.path.join(tuning_root, "weekly_*")) if os.path.isdir(p)]
    if not candidates:
        raise FileNotFoundError(f"No weekly output folder found under: {tuning_root}")

    dated_candidates = []
    for path in candidates:
        name = os.path.basename(path)
        m = re.fullmatch(r"weekly_(\d{8})", name)
        if m:
            dated_candidates.append((int(m.group(1)), path))

    if dated_candidates:
        dated_candidates.sort(key=lambda x: x[0], reverse=True)
        return dated_candidates[0][1]

    candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return candidates[0]


def _normalize(v: object) -> str:
    return str(v or "").strip()


def _is_missing_human_category(v: object) -> bool:
    text = _normalize(v).lower()
    return text in {"", "na", "n/a", "none", "null"}


def _valid_email(v: str) -> bool:
    text = _normalize(v).lower()
    return "@" in text and "." in text.split("@", 1)[-1]


def _safe_title(v: object, max_len: int = 140) -> str:
    text = _normalize(v).replace("\n", " ").replace("\r", " ")
    if len(text) <= max_len:
        return text
    return text[: max_len - 3] + "..."


def _pending_from_rows(rows: List[Dict[str, object]], source_label: str) -> Dict[str, List[Dict[str, str]]]:
    pending_by_assignee: Dict[str, List[Dict[str, str]]] = {}
    for row in rows:
        assignee = _normalize(row.get("assignee_email", "")).lower()
        if not _valid_email(assignee):
            continue
        if not _is_missing_human_category(row.get("human_category", "")):
            continue

        pending_by_assignee.setdefault(assignee, []).append(
            {
                "created_date": _normalize(row.get("created_date", "")),
                "technology": _normalize(row.get("technology", "")),
                "ips_title": _normalize(row.get("ips_title", "")),
            }
        )

    LOG.info("Loaded pending rows from %s: assignees=%d", source_label, len(pending_by_assignee))
    return pending_by_assignee


def _load_pending_from_csv(source_path: str) -> Dict[str, List[Dict[str, str]]]:
    with open(source_path, "r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        fields = {str(x).strip() for x in (reader.fieldnames or [])}
        required = {"assignee_email", "human_category", "ips_title"}
        missing = [x for x in required if x not in fields]
        if missing:
            raise RuntimeError(
                f"CSV missing required column(s): {missing}. Expected at least assignee_email, human_category, ips_title"
            )
        rows = [dict(row) for row in reader]
    return _pending_from_rows(rows, f"CSV {source_path}")


def _load_pending_from_xlsx(source_path: str) -> Dict[str, List[Dict[str, str]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError(f"openpyxl is required to read XLSX source: {exc}") from exc

    # Read via a temp copy to avoid transient OneDrive/Excel file locks.
    source_to_read = source_path
    temp_copy_path = ""
    try:
        fd, temp_copy_path = tempfile.mkstemp(prefix="labeling_source_", suffix=".xlsx")
        os.close(fd)
        shutil.copy2(source_path, temp_copy_path)
        source_to_read = temp_copy_path
    except Exception as exc:
        LOG.warning("Failed to create temp copy for XLSX (%s). Trying direct read.", exc)

    wb = load_workbook(source_to_read, read_only=True, data_only=True)
    try:
        ws = wb.active
        iterator = ws.iter_rows(values_only=True)
        header = next(iterator, None)
        if not header:
            return {}

        keys = [str(x or "").strip() for x in header]
        required = {"assignee_email", "human_category", "ips_title"}
        missing = [x for x in required if x not in set(keys)]
        # Defensive: detect generic "ColumnN" header row injected by some tools (e.g. pandas
        # re-export without header).  If row 1 looks like Column1/Column2/… but row 2 has real
        # headers, skip the spurious first row and re-read from row 2.
        if missing and all(k == "" or k.lower().startswith("column") for k in keys):
            next_row = next(iterator, None)
            if next_row:
                candidate_keys = [str(x or "").strip() for x in next_row]
                candidate_missing = [x for x in required if x not in set(candidate_keys)]
                if not candidate_missing:
                    LOG.warning(
                        "XLSX row 1 contained generic 'ColumnN' headers; treating row 2 as the real header row."
                    )
                    keys = candidate_keys
                    missing = candidate_missing
        if missing:
            raise RuntimeError(
                f"XLSX missing required column(s): {missing}. Expected at least assignee_email, human_category, ips_title"
            )

        rows: List[Dict[str, object]] = []
        for values in iterator:
            row_dict: Dict[str, object] = {}
            for idx, key in enumerate(keys):
                if not key:
                    continue
                row_dict[key] = values[idx] if idx < len(values) else ""
            rows.append(row_dict)

        return _pending_from_rows(rows, f"XLSX {source_path}")
    finally:
        wb.close()
        if temp_copy_path:
            try:
                os.remove(temp_copy_path)
            except Exception:
                pass


def _download_cloud_copy_if_possible(source_path: str, graph_auth_mode: str) -> str:
    """Try downloading a fresh OneDrive cloud copy for source-of-truth reads."""
    graph_path = _local_onedrive_to_graph_path(source_path)
    if not graph_path:
        return source_path

    mode = str(graph_auth_mode or os.getenv("GRAPH_AUTH_MODE", "delegated")).strip().lower()
    token = ""
    sender_upn = ""

    try:
        secret = resolve_graph_client_secret()
        if mode == "delegated":
            token = get_graph_token_delegated_with_secret(secret, scopes=["Files.Read"])
            url = f"{GRAPH_BASE}/me/drive/root:/{graph_path}:/content"
        else:
            sender_upn = str(os.getenv("GRAPH_SENDER_UPN") or "").strip()
            if not sender_upn:
                return source_path
            token = get_graph_token_app_only(secret, scopes=["https://graph.microsoft.com/.default"])
            sender_escaped = urllib.parse.quote(sender_upn, safe="")
            url = f"{GRAPH_BASE}/users/{sender_escaped}/drive/root:/{graph_path}:/content"

        req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"}, method="GET")
        with urllib.request.urlopen(req, timeout=30) as resp:
            payload = resp.read()

        suffix = os.path.splitext(source_path)[1] or ".tmp"
        fd, tmp_path = tempfile.mkstemp(prefix="labeling_cloud_", suffix=suffix)
        os.close(fd)
        with open(tmp_path, "wb") as fh:
            fh.write(payload)

        LOG.info("Using cloud-fresh source copy for pending check: %s", tmp_path)
        return tmp_path
    except Exception as exc:
        LOG.warning("Could not download cloud-fresh source; fallback to local file: %s", exc)
        return source_path


def _resolve_pending_source_path(csv_path: str) -> str:
    if csv_path.lower().endswith(".xlsx") and os.path.exists(csv_path):
        return csv_path

    xlsx_path = os.path.splitext(csv_path)[0] + ".xlsx"
    if os.path.exists(xlsx_path):
        # Team edits the shared XLSX link; treat XLSX as source of truth when available.
        return xlsx_path

    return csv_path


def _load_pending_by_assignee(source_path: str, csv_fallback_path: str = "") -> Dict[str, List[Dict[str, str]]]:
    if not os.path.exists(source_path):
        raise FileNotFoundError(f"Labeling source not found: {source_path}")

    if source_path.lower().endswith(".xlsx"):
        try:
            return _load_pending_from_xlsx(source_path)
        except Exception as exc:
            csv_fallback = os.path.splitext(source_path)[0] + ".csv"
            if os.path.exists(csv_fallback):
                LOG.warning("Failed to load XLSX source (%s); fallback to CSV: %s", exc, csv_fallback)
                return _load_pending_from_csv(csv_fallback)
            explicit_csv_fallback = str(csv_fallback_path or "").strip()
            if explicit_csv_fallback and os.path.exists(explicit_csv_fallback):
                LOG.warning(
                    "Failed to load XLSX source (%s); fallback to original CSV: %s",
                    exc,
                    explicit_csv_fallback,
                )
                return _load_pending_from_csv(explicit_csv_fallback)
            raise
    return _load_pending_from_csv(source_path)


def _display_name_from_email(email: str) -> str:
    local = _normalize(email).split("@", 1)[0]
    if not local:
        return "Engineer"
    return local.replace(".", " ").replace("-", " ").title()


def _local_onedrive_to_graph_path(path: str) -> str:
    norm = os.path.normpath(os.path.abspath(path))
    parts = [p for p in norm.split(os.sep) if p]
    one_drive_idx = -1
    for idx, part in enumerate(parts):
        if str(part).lower().startswith("onedrive - "):
            one_drive_idx = idx
            break
    if one_drive_idx < 0 or one_drive_idx + 1 >= len(parts):
        return ""
    rel_parts = parts[one_drive_idx + 1 :]
    return "/".join(urllib.parse.quote(p, safe="") for p in rel_parts)


def _build_csv_reference(
    *,
    csv_path: str,
    token: str,
    graph_auth_mode: str,
    sender_upn: str,
) -> str:
    graph_path = _local_onedrive_to_graph_path(csv_path)
    if not graph_path or not token:
        return csv_path

    mode = str(graph_auth_mode or "").strip().lower()
    if mode == "app" and sender_upn:
        sender_escaped = urllib.parse.quote(sender_upn, safe="")
        url = f"{GRAPH_BASE}/users/{sender_escaped}/drive/root:/{graph_path}:/createLink"
    else:
        url = f"{GRAPH_BASE}/me/drive/root:/{graph_path}:/createLink"

    # Use organization-wide editable link so assignees can update human_category directly.
    payload = {"type": "edit", "scope": "organization"}
    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    def _upload_to_onedrive_and_retry() -> str:
        if not os.path.exists(csv_path):
            return ""
        try:
            if mode == "app" and sender_upn:
                sender_escaped = urllib.parse.quote(sender_upn, safe="")
                put_url = f"{GRAPH_BASE}/users/{sender_escaped}/drive/root:/{graph_path}:/content"
            else:
                put_url = f"{GRAPH_BASE}/me/drive/root:/{graph_path}:/content"

            with open(csv_path, "rb") as fh:
                put_req = urllib.request.Request(
                    put_url,
                    data=fh.read(),
                    headers={
                        "Authorization": f"Bearer {token}",
                        "Content-Type": "application/octet-stream",
                    },
                    method="PUT",
                )
                urllib.request.urlopen(put_req, timeout=30).read()

            with urllib.request.urlopen(req, timeout=20) as resp2:
                data2 = json.loads(resp2.read().decode("utf-8", errors="ignore"))
                link2 = (((data2 or {}).get("link") or {}).get("webUrl") or "").strip()
                return link2
        except Exception as exc2:
            LOG.warning("Upload+createLink retry failed; fallback to local path: %s", exc2)
            return ""

    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            data = json.loads(resp.read().decode("utf-8", errors="ignore"))
            link = (((data or {}).get("link") or {}).get("webUrl") or "").strip()
            if link:
                return link
    except urllib.error.HTTPError as exc:
        if int(getattr(exc, "code", 0)) == 404:
            # New local file may not be synced to cloud yet; upload then retry.
            retried_link = _upload_to_onedrive_and_retry()
            if retried_link:
                return retried_link
        exc_text = str(exc)
        if "403" in exc_text:
            LOG.warning(
                "Failed to create shared link for CSV (403). Ensure delegated Graph permission Files.ReadWrite is granted/consented; fallback to local path. Details: %s",
                exc,
            )
        else:
            LOG.warning("Failed to create shared link for CSV; fallback to local path: %s", exc)
    except Exception as exc:
        exc_text = str(exc)
        if "403" in exc_text:
            LOG.warning(
                "Failed to create shared link for CSV (403). Ensure delegated Graph permission Files.ReadWrite is granted/consented; fallback to local path. Details: %s",
                exc,
            )
        else:
            LOG.warning("Failed to create shared link for CSV; fallback to local path: %s", exc)
    return csv_path


def _compose_body(email: str, rows: List[Dict[str, str]], csv_reference: str, max_items: int) -> str:
    now_text = datetime.now().strftime("%Y-%m-%d %H:%M")
    shown = rows[: max(1, max_items)]

    lines: List[str] = []
    lines.append(f"Hi {_display_name_from_email(email)},")
    lines.append("")
    lines.append("This is an automatic reminder for weekly labeling.")
    lines.append(f"You still have {len(rows)} assigned row(s) with empty human_category.")
    lines.append("")
    lines.append("ACTION REQUIRED:")
    lines.append("[需要處理]")
    lines.append("1. Open the source file below.")
    lines.append("1. 開啟下方來源檔案。")
    lines.append("2. Fill the HUMAN_CATEGORY column for your assigned rows.")
    lines.append("2. 請在你負責的列填寫 HUMAN_CATEGORY 欄位。")
    lines.append("3. Save after update.")
    lines.append("3. 更新後請儲存。")
    lines.append("4. Deadline: Friday 5:00 PM.")
    lines.append("4. 截止時間：本週五下午 5:00。")
    lines.append("")
    lines.append(f"Source file: {csv_reference}")
    lines.append(f"來源檔案: {csv_reference}")
    lines.append(f"Check time: {now_text}")
    lines.append(f"檢查時間: {now_text}")
    lines.append("")
    lines.append("Pending rows (sample):")
    lines.append("待處理列（節錄）：")
    for idx, row in enumerate(shown, start=1):
        created = _normalize(row.get("created_date", "")) or "-"
        tech = _normalize(row.get("technology", "")) or "-"
        title = _safe_title(row.get("ips_title", ""), max_len=160) or "(no title)"
        lines.append(f"{idx}. [{created}] [{tech}] {title}")

    if len(rows) > len(shown):
        lines.append(f"... and {len(rows) - len(shown)} more row(s).")

    lines.append("")
    lines.append("Please update HUMAN_CATEGORY as soon as possible. If already updated, please ignore this message.")
    lines.append("Deadline: Friday 5:00 PM.")
    lines.append("請盡快更新 HUMAN_CATEGORY；若已完成，請忽略此信。")
    lines.append("截止時間：本週五下午 5:00。")
    lines.append("Thanks.")
    lines.append("謝謝。")
    return "\n".join(lines)


def _safe_console_text(text: str) -> str:
    enc = sys.stdout.encoding or "utf-8"
    return text.encode(enc, errors="replace").decode(enc, errors="replace")


def _load_valid_categories(weight_map_path: str = "issue_category_weights.json") -> List[str]:
    """Return sorted list of valid human_category values from issue_category_weights.json."""
    try:
        with open(weight_map_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        normalized = {
            ("Need-Triage" if str(c) == "Needs-Triage" else str(c))
            for c in data.get("category_weights", {}).keys()
        }
        return sorted(normalized)
    except Exception as exc:
        LOG.warning("Could not load category weights for dropdown: %s", exc)
        return []


def _ensure_editable_workbook(source_path: str) -> str:
    """Return editable workbook path; create XLSX from CSV only when XLSX is missing.

    When creating a new XLSX, add a dropdown Data Validation on the
    human_category column so reviewers can only pick from valid categories.
    """
    if source_path.lower().endswith(".xlsx") and os.path.exists(source_path):
        return source_path

    xlsx_path = os.path.splitext(source_path)[0] + ".xlsx"
    if os.path.exists(xlsx_path):
        # Never overwrite existing team-edited workbook.
        return xlsx_path

    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore
    except Exception as exc:
        LOG.warning("openpyxl unavailable; keep using CSV reference (may open as read-only in web): %s", exc)
        return source_path

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "weekly_labeling_template"

        header_row: List[str] = []
        all_rows: List[List[str]] = []
        with open(source_path, "r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            for i, row in enumerate(reader):
                ws.append(row)
                if i == 0:
                    header_row = row
                else:
                    all_rows.append(row)

        # Add dropdown validation on human_category column
        if "human_category" in header_row:
            col_idx = header_row.index("human_category") + 1  # 1-based
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            valid_cats = _load_valid_categories(
                os.path.join(os.path.dirname(source_path), "..", "..", "issue_category_weights.json")
            ) or _load_valid_categories("issue_category_weights.json")

            if valid_cats:
                # Excel formula: "Cat1,Cat2,..." (max ~255 chars per DV formula)
                cat_formula = '"{}"'.format(",".join(valid_cats))
                dv = DataValidation(
                    type="list",
                    formula1=cat_formula,
                    allow_blank=True,
                    showDropDown=False,
                    showErrorMessage=True,
                    errorTitle="Invalid category",
                    error="Please select a category from the dropdown list.",
                )
                data_rows = len(all_rows)
                dv.sqref = f"{col_letter}2:{col_letter}{max(data_rows + 1, 200)}"
                ws.add_data_validation(dv)
                LOG.info("Added dropdown validation for %d categories on column %s", len(valid_cats), col_letter)

        wb.save(xlsx_path)
        return xlsx_path
    except Exception as exc:
        LOG.warning("Failed to create XLSX copy; keep using CSV reference: %s", exc)
        return source_path


def _send_reminders(
    *,
    pending_by_assignee: Dict[str, List[Dict[str, str]]],
    csv_path: str,
    cc_list: List[str],
    max_items: int,
    dry_run: bool,
    send_email: bool,
    graph_auth_mode: str,
    allow_email_failure: bool,
) -> int:
    recipients = sorted(pending_by_assignee.keys())
    if not recipients:
        LOG.info("No pending assignees found. No reminder email sent.")
        return 0

    total_rows = sum(len(v) for v in pending_by_assignee.values())
    LOG.info("Pending assignees: %d, pending rows: %d", len(recipients), total_rows)

    if dry_run:
        LOG.info("Dry-run mode enabled; emails will not be sent.")

    token = ""
    sender_upn = ""
    mode = str(graph_auth_mode or os.getenv("GRAPH_AUTH_MODE", "delegated")).strip().lower()

    if send_email and not dry_run:
        try:
            secret = resolve_graph_client_secret()
            if mode == "delegated":
                LOG.info("Graph auth mode: delegated")
                token = get_graph_token_delegated_with_secret(secret, scopes=["Mail.Send", "Files.ReadWrite"])
            else:
                sender_upn = str(os.getenv("GRAPH_SENDER_UPN") or "").strip()
                if not sender_upn:
                    raise RuntimeError("GRAPH_SENDER_UPN is required when GRAPH_AUTH_MODE=app.")
                LOG.info("Graph auth mode: app (sender=%s)", sender_upn)
                token = get_graph_token_app_only(secret, scopes=["https://graph.microsoft.com/.default"])
        except Exception as exc:
            if not allow_email_failure:
                raise
            LOG.warning("Failed to initialize Graph email; continuing because email failure is allowed: %s", exc)
            return 0

    sent_count = 0
    failed_count = 0
    reference_path = _ensure_editable_workbook(csv_path)
    csv_reference = _build_csv_reference(
        csv_path=reference_path,
        token=token,
        graph_auth_mode=mode,
        sender_upn=sender_upn,
    )

    for email in recipients:
        pending_rows = pending_by_assignee[email]
        subject = f"[Action Required][請處理] Fill HUMAN_CATEGORY / 請填寫 HUMAN_CATEGORY ({len(pending_rows)} pending) - Deadline Friday 5PM / 截止週五 5PM"
        body = _compose_body(email, pending_rows, csv_reference, max_items=max_items)

        if dry_run or not send_email:
            LOG.info("[DRY RUN] To=%s Cc=%s Subject=%s", email, ",".join(cc_list) if cc_list else "", subject)
            preview = body if len(body) <= 800 else body[:800] + "\n... (truncated)"
            print("\n--- Email preview ---")
            print(_safe_console_text(f"To: {email}"))
            if cc_list:
                print(_safe_console_text(f"Cc: {', '.join(cc_list)}"))
            print(_safe_console_text(f"Subject: {subject}"))
            print(_safe_console_text(preview))
            continue

        try:
            send_mail_via_graph(
                token=token,
                subject=subject,
                body_text=body,
                to_addrs=[email],
                cc_addrs=cc_list or None,
                save_to_sent_items=True,
                content_type="Text",
                sender_upn=sender_upn or None,
            )
        except Exception as exc:
            if not allow_email_failure:
                raise
            failed_count += 1
            LOG.warning("Failed to send reminder to %s; continuing because email failure is allowed: %s", email, exc)
            continue
        sent_count += 1
        LOG.info("Reminder sent to %s (pending=%d)", email, len(pending_rows))

    if dry_run or not send_email:
        LOG.info("Reminder dry-run completed.")
    else:
        LOG.info("Reminder email completed: sent=%d failed=%d allow_email_failure=%s.", sent_count, failed_count, allow_email_failure)
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Send reminder emails for missing human_category labels.")
    parser.add_argument("--tuning-root", default="tuning_outputs")
    parser.add_argument("--weekly-dir", default="", help="weekly_* directory containing weekly_labeling_template.csv")
    parser.add_argument("--csv", default="", help="Explicit path to weekly_labeling_template.csv")
    parser.add_argument("--recipients", default="recipients.json", help="Recipients JSON for default cc list")
    parser.add_argument("--cc", default="", help="Override cc list (comma-separated)")
    parser.add_argument("--max-items", type=int, default=15, help="Max pending rows shown in each email")
    parser.add_argument("--graph-auth-mode", choices=["app", "delegated"], default=os.getenv("GRAPH_AUTH_MODE", "delegated"))
    parser.add_argument("--send-email", dest="send_email", action="store_true", default=True)
    parser.add_argument("--no-email", dest="send_email", action="store_false")
    parser.add_argument(
        "--allow-email-failure",
        action="store_true",
        default=str(os.getenv("LABELING_REMINDER_ALLOW_EMAIL_FAILURE", "")).strip() == "1",
        help="Return success even if Graph reminder email sending fails.",
    )
    parser.add_argument("--dry-run-email", action="store_true")
    parser.add_argument("--log-level", default="INFO")
    args = parser.parse_args()

    _setup_logging(args.log_level)

    if args.csv:
        csv_path = os.path.abspath(str(args.csv).strip())
    else:
        weekly_dir = str(args.weekly_dir or "").strip()
        if not weekly_dir:
            weekly_dir = _find_latest_weekly_dir(str(args.tuning_root or "").strip())
        csv_path = os.path.abspath(os.path.join(weekly_dir, "weekly_labeling_template.csv"))

    source_path = _resolve_pending_source_path(csv_path)
    LOG.info("Pending source path: %s", source_path)

    to_list, cc_list = load_recipients(str(args.recipients or "").strip())
    _ = to_list
    if str(args.cc or "").strip():
        cc_list = [x.strip() for x in str(args.cc).split(",") if x.strip()]

    source_for_pending = source_path
    tmp_cloud_copy = ""
    if source_path.lower().endswith((".xlsx", ".csv")):
        source_for_pending = _download_cloud_copy_if_possible(source_path, str(args.graph_auth_mode or "delegated"))
        if source_for_pending != source_path:
            tmp_cloud_copy = source_for_pending

    try:
        pending_by_assignee = _load_pending_by_assignee(source_for_pending, csv_fallback_path=csv_path)
        return _send_reminders(
            pending_by_assignee=pending_by_assignee,
            csv_path=source_path,
            cc_list=cc_list,
            max_items=max(1, int(args.max_items)),
            dry_run=bool(args.dry_run_email),
            send_email=bool(args.send_email),
            graph_auth_mode=str(args.graph_auth_mode or "app"),
            allow_email_failure=bool(args.allow_email_failure),
        )
    finally:
        if tmp_cloud_copy:
            try:
                os.remove(tmp_cloud_copy)
            except Exception:
                pass


if __name__ == "__main__":
    raise SystemExit(main())
